"""XLSX workbook generation — 8-sheet styled workbook (ported from Tiger)."""

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import (
    CLASS_ORDER,
    CRITICAL,
    HIGH,
    OVERPASS_PRIMARY,
    SANITY_THRESHOLD,
)
from .zones import ZONES

log = logging.getLogger(__name__)

# --- Styles ---

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CRIT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
INDEX_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
MONO_FONT = Font(name="Courier New", size=10)
CRIT_FONT = Font(name="Arial", size=10, color="9C0006")
HIGH_FONT = Font(name="Arial", size=10, color="9C5700")
OK_FONT = Font(name="Arial", size=10, color="006100")

THIN_BORDER = Border(bottom=Side(border_style="thin", color="B4C6E7"))
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row_idx: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _style_data_cell(cell, font=DATA_FONT, align=DATA_ALIGN, border=THIN_BORDER) -> None:
    cell.font = font
    cell.alignment = align
    cell.border = border


def _autosize_columns(ws, max_widths: dict[int, int]) -> None:
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)


def _severity_style(cell, severity: str) -> None:
    if severity == CRITICAL:
        cell.font = CRIT_FONT
        cell.fill = CRIT_FILL
    elif severity == HIGH:
        cell.font = HIGH_FONT
        cell.fill = HIGH_FILL
    elif severity == "OK":
        cell.font = OK_FONT
        cell.fill = OK_FILL


def _write_row(ws, row_idx: int, values: list, widths: dict[int, int]) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_cell(cell)
        text_len = len(str(value)) if value is not None else 0
        widths[col_idx] = max(widths.get(col_idx, 0), text_len)


def write_xlsx(
    classified: dict,
    zone_key: str,
    out_path: Path,
    query_text: str,
    audit_ts: str,
    *,
    output_root: Path | None = None,
) -> None:
    """Generate the 8-sheet XLSX workbook."""
    zone = ZONES[zone_key]
    stats = classified["summary_stats"]
    wb = Workbook()

    # Sheet 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws.cell(row=1, column=1, value=f"TIGER/Line Audit - {zone['name']}").font = Font(
        name="Arial", size=14, bold=True, color="1F4E79"
    )
    ws.cell(row=1, column=2, value=f"Audit date (UTC): {audit_ts}").font = Font(
        name="Arial", size=10, italic=True
    )
    ws.cell(row=2, column=1, value=f"Source: Overpass API live query | Bbox (S,W,N,E): {zone['bbox']}").font = Font(
        name="Arial", size=10, italic=True, color="555555"
    )

    if stats.get("under_sanity_threshold"):
        warn = ws.cell(
            row=3, column=1,
            value=f"WARNING: Overpass returned fewer than {SANITY_THRESHOLD} elements. "
                  f"Audit may be based on truncated data.",
        )
        warn.font = Font(name="Arial", size=10, bold=True, color="9C0006")
        warn.fill = CRIT_FILL
        warn.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

    headers = ["Metric", "Count", "Significance"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    _style_header_row(ws, 4, len(headers))

    metrics = [
        ("Total unreviewed TIGER ways in zone", stats["total"],
         "All ways carrying tiger:reviewed=no in the bounding box."),
        ("Residential streets (unreviewed)", stats["residential"],
         "Subset where highway=residential."),
        ("Residential streets with oneway=yes (Class A)",
         sum(1 for w in classified["all_ways"] if w["highway"] == "residential" and w["oneway"] == "yes"),
         "Likely false one-way; routing engine sees no legal exit."),
        ("All road types with oneway=yes", stats["oneway_yes_total"],
         "Total one-way ways across all highway types."),
        ("Multi-segment disconnect risk streets (Class B)", stats["class_b_street_count"],
         "Named streets with 2+ unreviewed segments; intersection nodes may be disconnected."),
        ("Compound defect ways (Class AB)", stats["class_ab_count"],
         "False one-way AND on a multi-segment street; worst-case routing."),
    ]
    for i, (m, v, sig) in enumerate(metrics, start=5):
        ws.cell(row=i, column=1, value=m)
        c = ws.cell(row=i, column=2, value=v)
        c.number_format = "#,##0"
        ws.cell(row=i, column=3, value=sig)
        for col in (1, 2, 3):
            _style_data_cell(ws.cell(row=i, column=col))

    pct_row = 5 + len(metrics)
    ws.cell(row=pct_row, column=1, value="Residential % of total")
    pct = ws.cell(row=pct_row, column=2, value="=B6/B5")
    pct.number_format = "0.0%"
    ws.cell(row=pct_row, column=3, value="Share of unreviewed ways that are residential streets.")
    for col in (1, 2, 3):
        _style_data_cell(ws.cell(row=pct_row, column=col))

    ctx_start = pct_row + 2
    context_lines = [
        "CONTEXT",
        "SORTA operates MetroNow, a microtransit service in Hamilton County, Ohio,",
        "powered by Via Transportation. Via's routing engine consumes OpenStreetMap.",
        "The 2007-2008 TIGER/Line Census import seeded thousands of road segments",
        "with tiger:reviewed=no - meaning no human has verified them since import.",
        "False one-way tags and disconnected nodes inside this dataset translate",
        "directly into routing failures - service denials for transit-dependent riders.",
    ]
    for i, line in enumerate(context_lines):
        c = ws.cell(row=ctx_start + i, column=1, value=line)
        if i == 0:
            c.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
        else:
            c.font = Font(name="Arial", size=10, color="333333")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A5"

    # Sheet 2: Class A - Highest Risk (Class AB streets)
    ws2 = wb.create_sheet("Class A - Highest Risk")
    ab_by_street: dict[str, list[dict]] = defaultdict(list)
    for w in classified["class_ab"]:
        ab_by_street[w["name_display"]].append(w)
    ab_streets = sorted(ab_by_street.items(), key=lambda kv: kv[0].lower())
    headers2 = ["#", "Street", "Segments", "OSM Way IDs", "Severity", "Status"]
    for col, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, 1, len(headers2))
    widths2: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers2)}
    for idx, (street, ways) in enumerate(ab_streets, start=1):
        row = idx + 1
        way_ids = ", ".join(str(w["id"]) for w in ways)
        values = [idx, street, len(ways), way_ids, CRITICAL, "Not started"]
        _write_row(ws2, row, values, widths2)
        _severity_style(ws2.cell(row=row, column=5), CRITICAL)
    ws2.freeze_panes = "A2"
    _autosize_columns(ws2, widths2)

    # Sheet 3: Class A - Moderate Risk (Class A only, single-segment)
    ws3 = wb.create_sheet("Class A - Moderate Risk")
    a_only_sorted = sorted(
        classified["class_a_only"],
        key=lambda w: (w["name_display"].lower(), w["id"] or 0),
    )
    headers3 = ["#", "Street", "OSM Way ID", "Severity", "Status"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=h)
    _style_header_row(ws3, 1, len(headers3))
    widths3: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers3)}
    index_street = (zone.get("index-case-street") or "").strip()
    for idx, w in enumerate(a_only_sorted, start=1):
        row = idx + 1
        is_index = bool(index_street) and (w["name"] or "").strip() == index_street
        status = "INDEX CASE" if is_index else "Not started"
        values = [idx, w["name_display"], w["id"], CRITICAL, status]
        _write_row(ws3, row, values, widths3)
        _severity_style(ws3.cell(row=row, column=4), CRITICAL)
        if is_index:
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=col).fill = INDEX_FILL
    ws3.freeze_panes = "A2"
    _autosize_columns(ws3, widths3)

    # Sheet 4: Class B - Multi-Segment (streets with 5+ segments)
    ws4 = wb.create_sheet("Class B - Multi-Segment")
    headers4 = ["#", "Street", "Segments", "One-Way Segments", "Road Type", "Status"]
    for col, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=col, value=h)
    _style_header_row(ws4, 1, len(headers4))
    widths4: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers4)}
    big_b = [
        (street, ways)
        for street, ways in classified["class_b_streets"].items()
        if len(ways) >= 5
    ]
    big_b.sort(key=lambda kv: -len(kv[1]))
    for idx, (street, ways) in enumerate(big_b, start=1):
        row = idx + 1
        oneway_count = sum(1 for w in ways if w["oneway"] == "yes")
        types = sorted({w["highway"] or "(unset)" for w in ways})
        values = [idx, street, len(ways), oneway_count, ", ".join(types), "Not started"]
        _write_row(ws4, row, values, widths4)
        if oneway_count > 0:
            _severity_style(ws4.cell(row=row, column=4), CRITICAL)
    ws4.freeze_panes = "A2"
    _autosize_columns(ws4, widths4)

    # Sheet 5: All Ways
    ws5 = wb.create_sheet("All Ways")
    headers5 = [
        "Way ID", "Street Name", "Highway Type", "Oneway",
        "Defect Class", "Severity", "tiger:cfcc", "tiger:name_base",
        "Review Status", "Last Editor", "Last Edit",
    ]
    for col, h in enumerate(headers5, start=1):
        ws5.cell(row=1, column=col, value=h)
    _style_header_row(ws5, 1, len(headers5))
    widths5: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers5)}
    class_rank = {c: i for i, c in enumerate(CLASS_ORDER)}
    sorted_ways = sorted(
        classified["all_ways"],
        key=lambda w: (class_rank[w["defect_class"]], w["name_display"].lower(), w["id"] or 0),
    )
    for idx, w in enumerate(sorted_ways, start=2):
        values = [
            w["id"], w["name_display"], w["highway"] or "", w["oneway"] or "",
            w["defect_class"], w["severity"],
            w["tiger_cfcc"] or "", w["tiger_name_base"] or "",
            w.get("review_status", ""), w.get("user", ""), w.get("timestamp", ""),
        ]
        _write_row(ws5, idx, values, widths5)
        _severity_style(ws5.cell(row=idx, column=6), w["severity"])
    ws5.freeze_panes = "A2"
    _autosize_columns(ws5, widths5)

    # Sheet 6: MetroNow Zones
    ws6 = wb.create_sheet("MetroNow Zones")
    headers6 = ["Zone", "Audit Status", "Unreviewed Segments", "Notes"]
    for col, h in enumerate(headers6, start=1):
        ws6.cell(row=1, column=col, value=h)
    _style_header_row(ws6, 1, len(headers6))
    widths6: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers6)}
    for idx, (k, z) in enumerate(ZONES.items(), start=2):
        if k == zone_key:
            status = "AUDIT COMPLETE"
            seg: int | str = stats["total"]
            severity_for_row = "OK"
        else:
            other_csv = _find_other_zone_csv(k, output_root)
            if other_csv is not None:
                status = "AUDIT COMPLETE"
                seg = other_csv
                severity_for_row = "OK"
            else:
                status = "NOT STARTED"
                seg = ""
                severity_for_row = HIGH
        values = [z["name"], status, seg, z["description"]]
        _write_row(ws6, idx, values, widths6)
        _severity_style(ws6.cell(row=idx, column=2), severity_for_row)
    ws6.freeze_panes = "A2"
    _autosize_columns(ws6, widths6)

    # Sheet 7: Work Plan
    ws7 = wb.create_sheet("Work Plan")
    headers7 = ["Phase", "Description", "Items", "Min/Item", "Hours"]
    for col, h in enumerate(headers7, start=1):
        ws7.cell(row=1, column=col, value=h)
    _style_header_row(ws7, 1, len(headers7))
    widths7: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers7)}
    p4_items = max(stats["residential"] - stats["class_a_count"], 0)
    work_rows = [
        ("P1", "Fix compound defects (Class AB)", stats["class_ab_count"], 12),
        ("P2", "Fix single-segment false one-ways (Class A only)", stats["class_a_only_count"], 4),
        ("P3", "Inspect multi-segment disconnects (Class B)", stats["class_b_street_count"], 7),
        ("P4", "Review remaining residential ways", p4_items, 2),
    ]
    for i, (phase, desc, items, mins) in enumerate(work_rows, start=2):
        ws7.cell(row=i, column=1, value=phase)
        ws7.cell(row=i, column=2, value=desc)
        ws7.cell(row=i, column=3, value=items).number_format = "#,##0"
        ws7.cell(row=i, column=4, value=mins).number_format = "#,##0"
        h_cell = ws7.cell(row=i, column=5, value=f"=C{i}*D{i}/60")
        h_cell.number_format = "#,##0.0"
        for col in range(1, len(headers7) + 1):
            _style_data_cell(ws7.cell(row=i, column=col))
            widths7[col] = max(widths7.get(col, 0), len(str(ws7.cell(row=i, column=col).value)))
    total_row = len(work_rows) + 2
    ws7.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
    ws7.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row - 1})").number_format = "#,##0.0"
    for col in (1, 5):
        _style_data_cell(ws7.cell(row=total_row, column=col))
    ws7.freeze_panes = "A2"
    _autosize_columns(ws7, widths7)

    # Sheet 8: Overpass Query
    ws8 = wb.create_sheet("Overpass Query")
    info_lines = [
        ("Endpoint", OVERPASS_PRIMARY),
        ("Bounding box (S, W, N, E)", str(zone["bbox"])),
        ("Executed (UTC)", audit_ts),
        ("", ""),
        ("Query", ""),
    ]
    for i, (k, v) in enumerate(info_lines, start=1):
        kc = ws8.cell(row=i, column=1, value=k)
        vc = ws8.cell(row=i, column=2, value=v)
        kc.font = Font(name="Arial", size=10, bold=True)
        vc.font = MONO_FONT
    base = len(info_lines) + 1
    for i, line in enumerate(query_text.splitlines()):
        c = ws8.cell(row=base + i, column=1, value=line)
        c.font = MONO_FONT
    ws8.column_dimensions["A"].width = 60
    ws8.column_dimensions["B"].width = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("XLSX saved: %s", out_path)


def _find_other_zone_csv(zone_key: str, output_root: Path | None) -> int | None:
    """Check if another zone's all_ways.csv exists and return its row count."""
    if output_root is None:
        return None
    csv_path = output_root / f"osm-audit-{zone_key}" / "csv" / "all-ways.csv"
    if not csv_path.exists():
        return None
    try:
        with csv_path.open("r", encoding="utf-8") as fh:
            return max(sum(1 for _ in fh) - 1, 0)
    except OSError:
        return None
